import tkinter as tk
from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
import matplotlib.pyplot as plt
import mysql.connector as sql
from tkinter.simpledialog import askstring
from functools import reduce
import smtplib
import re
import time
from pymongo import MongoClient
import turtle
import random
import xlsxwriter
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter,inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, Table



count = 1
backcount = 0
accounts = []



class Login():

	def __init__(self):
		self.email = ""
		self.name = ""
		self.line = None
		self.keep_logged_in = 0

	def drive_ret(self):
		global accounts
		accounts = list(map(lambda item: list(item.values()), records.find()))
		

	def sign_in_check(self):
		global accounts
		self.drive_ret()

		flag = False

		if (accounts != None and accounts != []):
			with open("logged_in.txt","r") as emailfile:
				eml=emailfile.read()

			for i in range(0, len(accounts), 1):

				if (accounts[i][1] == eml):
					flag = True
					self.email = accounts[i][1]
					self.line = i

					self.name = accounts[i][2]

		if (accounts == None or accounts == False or flag == False):
			self.page1()

	def Forgot_Password(self):

		def part2():

			global win5
			win5 = tk.Tk()
			#            win5.iconbitmap('Logo.ico')
			win5.title("Account recovery")
			win5.geometry("200x200")
			win5.config(bg="Turquoise")
			Label(win5, text="Enter the new password").place(x=40, y=10)
			global ent3
			ent3 = Entry(win5, width=15)
			ent3.place(x=40, y=50)
			tk.Button(win5, text="Submit", command=lambda: getem(2)).place(x=70, y=100)
			win5.mainloop()

		def getem(step):
			global accounts

			if (step == 1):
				global var1, var2
				s = smtplib.SMTP('smtp.gmail.com', 587)
				global count
				if (count == 1):
					count += 1
					var1 = ent1.get()
					s.starttls()
					s.login("instafoodalp@gmail.com", "alpine2003$")

					message = f"Account Recovery\nYour otp is {otp}"
					s.sendmail("instafoodalp@gmail.com", var1, message)
					s.quit()
					s.close()
					tk.Button(win4, text="Submit", padx=10, pady=10, command=lambda: getem(1)).place(x=220, y=100)
				elif (count == 2):
					var2 = ent2.get()
					var2 = var2.strip(" ")

					if (str(otp) != var2):

						win4.destroy()
						self.Forgot_Password()
					else:

						win4.destroy()
						part2()

			elif (step == 2):
				global var3

				var3 = ent3.get()
				if (accounts != None and accounts != False):
					records.update_one({"email": var1}, {'$set': {'password': var3}})
					accounts = list(map(lambda item: list(item.values()), records.find()))
				else:
					r=Tk()
					r.withdraw()
					messagebox.showerror("Pop-Up","No account exist.")
					r.destroy()
					self.page1()
				win5.destroy()

		count = 1
		win4 = tk.Tk()
		win4.title("Account recovery")
		win4.iconbitmap('Logo.ico')
		win4.geometry("320x170")
		win4.config(bg="Turquoise1")
		otp = random.randint(1000, 9999)

		Label(win4, text="Enter the email . Otp will be sent to your email address", font=("Times New Roman", 10),
			  bg="Turquoise1").place(x=10, y=10)
		Label(win4, text="Email").place(x=15, y=40)
		Label(win4, text="Enter Otp").place(x=15, y=80)

		ent1 = Entry(win4, width=30)
		ent1.place(x=80, y=40)

		ent2 = Entry(win4, width=10)
		ent2.place(x=80, y=80)
		Button(win4, text="Submit", padx=10, pady=10, command=lambda: getem(1)).place(x=220, y=100)

		win4.mainloop()

	def signin_up(self, case, page):
		global win2
		win2 = tk.Toplevel()
		win2.iconbitmap('Logo.ico')
		if (page == 1):
			var = StringVar()
			var.set("Email")
			sct = Image.open("scooter.jpg")
			sct = sct.resize((250, 250))
			sctphoto = ImageTk.PhotoImage(sct)
			sctlabel = tk.Label(win2, image=sctphoto)
			sctlabel.place(x=0, y=0)

		elif (page == 2):
			var = StringVar()
			var.set(log.email)

		win2.title("Instafood ->> Sign Up")
		win2.geometry("250x255")

		if (case == 1 and page == 1):
			Button(win2, text="Forgot Password?", bg="pink1", command=lambda: self.Forgot_Password()).place(
				x=15, y=155)

		tk.Label(win2, textvariable=var).place(x=10, y=30)
		Label(win2, text="Password").place(x=10, y=70)

		global n2, a2, p2

		if (page == 1):
			global e2, checkbox, selected_option

			selected_option = IntVar()
			checkbox = tk.Checkbutton(win2, onvalue=1, offvalue=0, bg="pink1", text="Keep me signed in ",variable=selected_option)
			checkbox.place(x=15, y=195)
			e2 = tk.Entry(win2)
			e2.place(x=70, y=30)
			Button(win2, text="\nSubmit\n", bg="pink1", command=lambda: self.collectdata(case)).place(x=180,y=160)
		elif (page == 2):
			Button(win2, text="\nSubmit\n", bg="pink1", command=lambda: home.collect()).place(x=180, y=160)

		n2 = tk.Entry(win2)
		a2 = tk.Entry(win2)

		if (case == 2):
			Label(win2, text="Name").place(x=10, y=110)

			Label(win2, text="Pincode").place(x=10, y=150)

			n2.place(x=70, y=110)
			a2.place(x=70, y=150)

		p2 = tk.Entry(win2)

		p2.place(x=70, y=70)

		win2.mainloop()

	def collectdata(self, num):
		global li, n
		n = num
		l = [e2, p2, n2, a2]
		li = [l[i].get() for i in range(0, 2 * num)]
		li.append(selected_option.get())

		self.email = e2.get()
		self.name = n2.get()
		self.keep_logged_in = selected_option.get()

		win2.destroy()
		win1.destroy()

	def page1(self):
		global win1
		win1 = tk.Tk()
		win1.iconbitmap('Logo.ico')
		win1.geometry("500x430")
		win1.title("Instafood.com")

		back = Image.open("food.png").resize((500, 420))
		backphoto = ImageTk.PhotoImage(back)
		background_img = Label(win1, image=backphoto)
		background_img.place(x=0, y=0)

		Welcome_label = Label(win1, text="Welcome!", font=('Lucida Calligraphy', 30), bg="black", fg="white")
		Welcome_label.place(x=30, y=30)

		logo = Image.open("instafood.png")
		logo = logo.resize((200, 250))
		logophoto = ImageTk.PhotoImage(logo)
		logolabel = Label(win1, image=logophoto)
		logolabel.place(x=250, y=80)

		Button(win1, text="\n\tSign In\t\t\n", command=lambda: self.signin_up(1, 1)).place(x=50, y=130)
		Button(win1, text="\n\tSign Up\t\t\n", command=lambda: self.signin_up(2, 1)).place(x=50, y=230)

		win1.mainloop()

		result = re.match('[0-9a-z]+@[a-z]+\.[a-z]+', li[0])

		if (result == None):
			r=Tk()
			r.withdraw()
			messagebox.showerror("Pop-Up","Email-Id entered is In-valid.")
			r.destroy()
			self.page1()
		else:
			li[0] = li[0].strip('\n')
			self.emailprocess(li, n)

	def emailprocess(self, details, cond):
		global accounts

		if (cond == 1):

			flag = False
			for i in range(0, len(accounts), 1):
				if (accounts[i][1] == details[0]):
					if (accounts[i][2] == details[1]):
						self.line = i
						flag = True

			if (flag == True and self.keep_logged_in == 1):
				with open("logged_in.txt","w") as emailfile:
					eml=emailfile.write(self.email)

			if (flag == False):
				r=Tk()
				r.withdraw()
				messagebox.showerror("Pop-Up","No account created with this EmailID.Please Sign Up to continue or Re-enter the Email-Id and password.")
				r.destroy()
				self.page1()

		if (cond == 2):
			flag = False

			if (accounts != False and accounts != None):

				for i in range(0, len(accounts), 1):
					if (accounts[i][1] == details[0]):
						flag = True
						r=Tk()
						r.withdraw()
						messagebox.showerror("Pop-Up","An account already exists with this Email Id. Please sign in to contiue. Click Forgot Password to retrieve the password.")
						r.destroy()
						self.page1()

			if (accounts == False or flag == False or accounts == [[]]):
				with open("logged_in.txt","w") as emailfile:
					if(self.keep_logged_in==1):
						eml=emailfile.write(self.email)
					else:
						pass
				fields = ["email", 'password', "name", "address"]
				new_user = {value: details[i] for i, value in enumerate(fields)}
				new_user["orders"]=[]
				records.insert_one(new_user)


class Homepage(Login):
	global cart
	global start
	global currentpos
	cart = {}
	start = 0
	currentpos = 0

	def suggestions(self):
		global framesug
		def prevord(ordli,num):
			x=num
			for i in range(len(ordli)-1+num,-1,-1):
				if(ordli[i][0]!=ordli[x][0]):
					num=i
					break
			lis.append(-(len(ordli)-num))
			return -(len(ordli)-num)

		try:
			dic=records.find_one({"email": log.email})
			ordli=dic["orders"]
			Label(framesug,bg="lightyellow",text="Based on your previous orders", fg="SlateBlue2", font=("Times New Roman", 15)).place(x=110,y=0)
			Label(framesug,bg="lightyellow",text="Food\t\t   Restaurant\t\tPrice").place(x=80,y=30)
			
			num=-1
			lis=[-1]

			Label(framesug,bg="lightyellow",text=str(ordli[num][0]+"\t  "+ordli[num][1]+"\t\t"+str(ordli[num][2]))).place(x=80,y=50)
			Button(framesug,command=lambda:self.orderagain(lis[0]),text="Order1",bg="Orange").place(x=10,y=50)
			num=prevord(ordli,num)

			Label(framesug,bg="lightyellow",text=str(ordli[num][0]+"\t  "+ordli[num][1]+"\t\t"+str(ordli[num][2]))).place(x=80,y=90)
			Button(framesug,command=lambda:self.orderagain(lis[1]),text="Order2",bg="Orange").place(x=10,y=90)
			num=prevord(ordli,num)

			Label(framesug,bg="lightyellow",text=str(ordli[num][0]+"\t  "+ordli[num][1]+"\t\t"+str(ordli[num][2]))).place(x=80,y=130)
			Button(framesug,command=lambda:self.orderagain(lis[2]),text="Order3",bg="Orange").place(x=10,y=130)
		except:
			pass

	def orderagain(self,ordnum):

		ordlist=accounts[log.line][5][ordnum]
		FoodName=ordlist[0]
		rest=ordlist[1]
		cusine=ordlist[4]
		price=ordlist[2]
		qty = 1
		r = len(rest)
		f = len(FoodName)

		FoodId = cusine[0:3] + FoodName[0:2] + FoodName[f - 2:f] + rest[0:2] + rest[r - 2:r]
		if(FoodId not in cart):
			cart[FoodId] = (FoodName, price, qty)
		elif(cart[FoodId][2]<7):
			cart[FoodId] = (FoodName, price, qty+1)
		#FoodName, price, rest, cus


	def addorder(self,ordli):
		global accounts
		try:
			accounts[log.line][5].append(ordli)
			li=accounts[log.line][5]
		except:
			accounts[0][5].append(ordli)
			li=accounts[0][5]

		
		orde={"orders": li }

		records.update_one({"email":log.email}, {'$set': orde})

	def statistics(self):
		global accounts
		sizes=[0,0,0,0,0,0,0,0]
		labels = ["Chinese", "South Indian", "North Indian", "Desserts", "Beverages", "Fast Food", "Snacks", "Italian"]	
		try:
			for i,acclist in enumerate(accounts):
				for j in acclist[5]:
					cus=j[4]
					sizes[labels.index(cus)]+=1
		except :
			pass

		add=reduce(lambda acc,item: acc+item ,sizes)

		size=[(float(i/add)*100) for i in sizes]

		sizelables = [str(str(val)+"%") for val in size]

		colors = ['yellowgreen', 'gold', 'lightskyblue', 'lightcoral',"yellow","pink","orange","palegreen"]
		text = []
		patch, text = plt.pie(sizes, labels=sizelables, colors=colors, startangle=90)
		plt.axis('equal')
		plt.legend(patch, labels, loc=(0,0))

		plt.show()

	def order_page(self):
		def get_data():
			global data
			data = variable.get()
			variable.set("Home")
			if (data == "Edit Profile"):
				self.edit_profile()
			elif (data == "Sign Out"):
				self.Sign_Out()
			elif (data == "Track My order"):
				trans.time_calc()
				trans.tracker()
			elif (data == "Home"):
				pass
			elif (data == "Popularity"):
				self.statistics()
			elif (data == "About Us"):
				self.info()

		global backcount

		global win3
		if (backcount == 0):
			win3 = tk.Tk()
			win3.iconbitmap('Logo.ico')
			win3.geometry("500x550")
			win3.title("Instafood >> Order")
			win3.configure(bg="pale green")
		backcount += 1
		global frame1
		frame1 = Frame(win3, bg="lightyellow", width=500, height=100)
		frame1.pack(side=LEFT)
		frame1.place(x=0, y=0)
		Label(frame1, text="ORDER", bg="lightyellow", fg="SlateBlue2", font=("Times New Roman", 20)).place(x=200, y=0)

		Label(frame1, text="Options", padx=15, pady=5, bg="turquoise1").place(x=5, y=5)
		OptionList = ["Home", "My Orders", "Track My order", "Popularity", "Edit Profile", "Sign Out", "About Us"]
		variable = tk.StringVar(frame1)
		variable.set("Home")
		opt = tk.OptionMenu(frame1, variable, *OptionList)
		opt.config(width=15, bg="turquoise1")
		opt.place(x=5, y=50)
		Button(frame1, text="click me to perform action", bg="orange", command=lambda: get_data()).place(x=150, y=52)

		logo = Image.open("Logo.ico")
		logo = logo.resize((120, 120))
		logophoto = ImageTk.PhotoImage(logo)
		logolabel = Label(frame1, image=logophoto, bg="lightyellow")
		logolabel.place(x=350, y=0)

		global framesug
		framesug=Frame(win3,bg="lightyellow",width=500,height=200)
		framesug.place(x=0,y=390)

		self.suggestions()

		self.FoodData()
		self.Database()
		win3.mainloop()

	def edit_profile(self):

		self.signin_up(2, 2)

	def collect(self):
		global p2, n2, a2

		upd = {'password': p2.get(), "name": n2.get(), "address": a2.get()}
		records.update_one({"email": log.email}, {'$set': upd})
		global win2
		win2.destroy()

	def Sign_Out(self):
		with open("logged_in.txt","w") as emailfile:
			pass
		win3.destroy()
                          
	def info(self):

		top = Tk()  
		top.geometry("515x343")
		var = StringVar()  
		msg = Message( top, text ="""About Us:\nWhat is InstaFood's Story? How did it all begin?\nAn avocational project created and excecuted with the help of three main individuals Achintya,Allen and Surya.\nInstafood is an online food ordering and delivery platform currently serving in India.
			\nThe innovative technology, large and nimble delivery service,and exceptional customer focus at Instafood enables a host of benefits that includes lightning fast deliveries, live order tracking and no limit on order amount, 
			all while having the pleasure of enjoying your favourite meal wherever you'd like it from wherever you want.\n The Instafood project was initiated on May 2020 to ensure no one goes hungry during the lockdown.
			\nAs we progress through time, we keenly look forward to spread our horizons with your help.\nThank You for all your help:)""",relief=SUNKEN,bg="black",fg="white",font="Courier")  
		msg.pack()  
		top.mainloop()  

	def FoodData(self):
		foodlist = [['Paneer Dosa', 'South Indian', 'Dosa and Idly', 'Bangalore Dosa Co.', '125', ''],
					['Onion Chilli Cheese Dosa', 'South Indian', 'Dosa and Idly', 'Bangalore Dosa Co.', '125', ''],
					['Spinach Corn Cheese Dosa', 'South Indian', 'Dosa and Idly', 'Bangalore Dosa Co.', '150', ''],
					['Andhra Chilli Potato Fry Dosa', 'South Indian', 'Dosa and Idly', 'Bangalore Dosa Co.', '125', ''],
					['Gulab Jamun', 'Desserts', 'Indian', 'Bangalore Dosa Co.', '50', ''],
					['Panneer Paratha', 'North Indian', 'Paratha', 'Paratha Envy', '143', ''],
					['Rabdii', 'Desserts', 'Indian', 'Paratha Envy', '40', ''],
					['Aloo Paratha Combo', 'North Indian', 'Paratha', 'Paratha Envy', '140', ''],
					['Chicken Manchurian', 'Chinese', 'Starters', 'BeijingBites', '240', 'FALSE'],
					['Veg Schezwan Fried Rice', 'Chinese', 'Fried Rice', 'BeijingBites', '150', ''],
					['Hong Kong Chicken', 'Chinese', 'Main Course', 'BeijingBites', '200', 'FALSE'],
					['Margherita-Small', 'Italian', 'Pizza', 'Pizza Hut', '120', ''],
					['Choco Volcano Treat Pack-2pcs', 'Desserts', '', 'Pizza Hut', '100', ''],
					['Pepsi Pet Bottle', 'Beverages', 'Soft Drink', 'Pizza Hut', '60', ''],
					['Spanish Tomato Veg Pasta', 'Italian', 'Pasta', 'Pizza Hut', '170', ''],
					['Chicken Tikka (Personal)', 'Italian', 'Pizza', 'Pizza Hut', '370', 'FALSE'],
					['Triple Chicken Feast New (Personal)', 'Italian', 'Pizza', 'Pizza Hut', '400', 'FALSE'],
					['Garlic Breadstix', 'Fast Food', '', 'Pizza Hut', '100', ''],
					['Idli (2pcs)', 'South Indian', 'Dosa and Idly', 'Sagar Fast Food', '24', ''],
					['Vada(1pc)', 'South Indian', 'South Indian Specialites', 'Sagar Fast Food', '24', ''],
					['Rava Idly', 'South Indian', 'Dosa and Idly', 'Sagar Fast Food', '32', ''],
					['Masala Dosa', 'South Indian', 'Dosa and Idly', 'Sagar Fast Food', '50', ''],
					['Poori Sagu', 'South Indian', 'South Indian Specialites', 'Sagar Fast Food', '40', ''],
					['Rava Onion Masala Dosa', 'South Indian', 'South Indian Specialites', 'Sagar Fast Food', '70', ''],
					['Paper Masala Dosa', 'South Indian', 'South Indian Specialites', 'Sagar Fast Food', '65', ''],
					['Khara Bath', 'South Indian', 'South Indian Specialites', 'Sagar Fast Food', '25', ''],
					['Schezwan Fried Rice', 'Chinese', 'Fried Rice', 'Sagar Fast Food', '100', ''],
					['Pakora', 'Snacks', '', 'Sagar Fast Food', '30', ''],
					['Butter Chicken', 'Fast Food', 'Rolls', 'Rolls on wheels', '130', 'FALSE'],
					['Butter Paneer Rice', 'North Indian', 'Rice', 'Rolls on wheels', '140', ''],
					['Schezwan Babycorn', 'Fast Food', 'Rolls', 'Rolls on wheels', '110', ''],
					['Schezwan Chicken', 'Fast Food', 'Rolls', 'Rolls on wheels', '139', 'FALSE'],
					['Mixed Veg Mayo', 'Fast Food', 'Rolls', 'Rolls on wheels', '100', ''],
					['Chicken Tava Fried', 'Fast Food', 'Rolls', 'Rolls on wheels', '130', 'FALSE'],
					['Chicken Tikka', 'Fast Food', 'Rolls', 'Rolls on wheels', '140', 'FALSE'],
					['Chicken Kaati Kebab', 'Fast Food', 'Rolls', 'Rolls on wheels', '140', 'FALSE'],
					['Aloo Tikiya', 'Fast Food', 'Rolls', 'Rolls on wheels', '100', ''],
					['Egg Roll ( 2 Eggs )', 'Fast Food', 'Rolls', 'Rolls on wheels', '90', 'FALSE'],
					['Vegetable Pulao', 'North Indian', 'Rice', 'Rolls on wheels', '140', ''],
					['Red Velvet Choco Sin', 'Desserts', 'Ice Cream', 'Polar Bear', '190', ''],
					['Cookie Monster Sundae', 'Desserts', 'Ice Cream', 'Polar Bear', '190', ''],
					['Death By Chocolate', 'Desserts', 'Ice Cream', 'Polar Bear', '220', ''],
					['Lychee Sundae', 'Desserts', 'Ice Cream', 'Polar Bear', '190', ''],
					['Purple Punch', 'Desserts', 'Ice Cream', 'Polar Bear', '100', ''],
					['Choco Almond Fudge', 'Desserts', 'Ice Cream', 'Polar Bear', '100', ''],
					['Belgian Chocolate', 'Desserts', 'Ice Cream', 'Polar Bear', '100', ''],
					['Alphonso Mango', 'Desserts', 'Ice Cream', 'Polar Bear', '90', ''],
					['Fig O Honey', 'Desserts', 'Ice Cream', 'Polar Bear', '90', ''],
					['Tender Coconut', 'Desserts', 'Ice Cream', 'Polar Bear', '100', ''],
					['Chocolate Milkshake', 'Beverages', 'Milkshake', 'Polar Bear', '150', ''],
					['Chicken Whopper', 'Fast Food', 'Burger', 'Burger King', '160', 'FALSE'],
					['Chicken Chilli Cheese Melt', 'Fast Food', 'Burger', 'Burger King', '130', 'FALSE'],
					['BK Veggie', 'Fast Food', 'Burger', 'Burger King', '90', ''],
					['Crispy Veg Supreme', 'Fast Food', 'Burger', 'Burger King', '70', ''],
					['Veg Whopper', 'Fast Food', 'Burger', 'Burger King', '150', ''],
					['Cold Coffee', 'Beverages', '', 'Burger King', '100', ''],
					['Medium Fries', 'Fast Food', 'Fries', 'Burger King', '80', ''],
					['Chicken Wings 15 Pieces', 'Fast Food', '', 'Burger King', '400', 'FALSE'],
					['Burritos', 'Mexican', 'Appetizers', 'Tacos Mexico', '210', ''],
					['Guacamole', 'Mexican', 'Appetizers', 'Tacos Mexico', '180', ''],
					['Quesadilas', 'Mexican', 'Appetizers', 'Tacos Mexico', '160', ''],
					['Tacos', 'Mexican', 'Appetizers', 'Tacos Mexico', '120', ''],
					['Nachos with spiced salsa', 'Mexican', 'Appetizers', 'Tacos Mexico', '200', ''],
					['Spaghetti', 'Italian', 'Appetizers', 'Little Italy', '200', ''],
					['Tirmisu', 'Italian', 'Appetizers', 'Little Italy', '350', ''],
					['Ravioli', 'Italian', 'Appetizers', 'Little Italy', '220', ''],
					['Penne Pasta', 'Italian', 'Appetizers', 'Little Italy', '180', ''],
					['Lasagna', 'Italian', 'Appetizers', 'Little Italy', '300', ''],
					['Pani puri', 'Snacks', 'Chaats', 'Chaat Corner', '50', ''],
					['Aloo tikki chole', 'Snacks', 'Chaats', 'Chaat Corner', '90', ''],
					['Bhel puri', 'Snacks', 'Chaats', 'Chaat Corner', '65', ''],
					['Dahi papdi', 'Snacks', 'Chaats', 'Chaat Corner', '30', ''],
					['Samosa', 'Snacks', 'Chaats', 'Chaat Corner', '20', ''],
					['Sev puri', 'Snacks', 'Chaats', 'Chaat Corner', '40', ''],
					['Dahi puri', 'Snacks', 'Chaats', 'Chaat Corner', '50', ''],
					['Papdi', 'Snacks', 'Chaats', 'Chaat Corner', '15', ''],
					['Dahi vada', 'Snacks', 'Chaats', 'Chaat Corner', '20', ''],
					['Plain Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '30', ''],
					['Mysore Masala dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '40', ''],
					['Set Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '35', ''],
					['Neer Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '35', ''],
					['Atta Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '40', ''],
					['Egg Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '50', ''],
					['Ragi Dosa', 'South Indian', 'Dosa and Idly', 'Dosa Corner', '40', ''],
					['Grilled Sandwich', 'Fast Food', 'Sandwiches', 'Food Plaza', '120', ''],
					['Paneer Sandwich', 'Fast Food', 'Sandwiches', 'Food Plaza', '110', ''],
					['Grilled Cheese Sandwich', 'Fast Food', 'Sandwiches', 'Food Plaza', '130', ''],
					['Veg Cheese Sandwich', 'Fast Food', 'Sandwiches', 'Food Plaza', '100', ''],
					['Mayo Veggie Sandwich', 'Fast Food', 'Sandwiches', 'Food Plaza', '110', ''],
					['Pav Bhaji', 'Snacks', 'Chaats', 'Chai Point', '140', ''],
					['Kashmiri Kahwa Chai UniFlask', 'Beverages', 'Hot Drink', 'Chai Point', '100', ''],
					['Ginger lemon Chai UniFlask', 'Beverages', 'Hot Drink', 'Chai Point', '90', ''],
					['Sulemani Chai UniFlask', 'Beverages', 'Hot Drink', 'Chai Point', '90', ''],
					['Mango Shake', 'Beverages', 'Milkshake', 'Chai Point', '120', ''],
					['Millet Pongal', 'South Indian', 'South Indian Specialites', 'Chai Point', '100', ''],
					['Anda Paratha with Cheese', 'North Indian', 'Paratha', 'Chai Point', '110', 'FALSE'],
					['Kesar Badam Lassi', 'Beverages', 'Lassi', 'Chai Point', '120', ''],
					['Chicken Shawarma', 'Middle Eastern', 'Shawarma', 'Arabian Hut', '100', 'FALSE'],
					['Grilled kebab shawarma', 'Middle Eastern', 'Shawarma', 'Arabian Hut', '120', 'FALSE'],
					['Paneer Shawarma', 'Middle Eastern', 'Shawarma', 'Arabian Hut', '130', ''],
					['Ratatouille', 'French', 'Appetizers', 'Hotel France', '225', ''],
					['Lobster bisque', 'French', 'Appetizers', 'Hotel France', '300', 'FALSE'],
					['Spinach soufflŠ', 'French', 'Appetizers', 'Hotel France', '250', ''],
					['CrŠpes', 'French', 'Appetizers', 'Hotel France', '175', ''],
					['Cheesy Croissant Casserole', 'French', 'Appetizers', 'Hotel France', '200', ''],
					['French Apple Tart', 'French', 'Appetizers', 'Hotel France', '200', '']]
		f=open("Pass.txt","r")
		pas=f.read()
		pas=pas[0:len(pas)-1]
		mydb = sql.connect(host="localhost", user="root",passwd=pas,charset="utf8")
		mycursor = mydb.cursor()
		try:
			mycursor.execute("CREATE Database Instafood")
		except:
			pass
		finally:
			mycursor.execute("USE Instafood")

		try:
			mycursor.execute(
				"CREATE TABLE Food(FoodName char(45),CusineOfFood char(20),Type char(40),Restaurant char(40),Price INT,Veg BOOLEAN DEFAULT TRUE)")
			mycursor.execute("DROP TABLE Food")
			mycursor.execute(
				"CREATE TABLE Food(FoodName char(45),CusineOfFood char(20),Type char(40),Restaurant char(40),Price INT,Veg BOOLEAN DEFAULT TRUE)")
			for i in foodlist:
				if (i[5] == ''):
					i[5] = 1
				else:
					i[5] = 0
				i[4] = int(i[4])

				tu = tuple(i)

				s = "INSERT INTO Food(FoodName,CusineOfFood,Type,Restaurant,Price,Veg) VALUES (%s,%s,%s,%s,%s,%s)"
				mycursor.execute(s, tu)
		except:
			pass
		finally:
			mydb.commit()

	def Database(self):
		global frame2
		frame2 = Frame(win3, width=500, height=230, bg="pale green")
		frame2.place(x=0, y=100)

		# Establishing connections with database
		f=open("Pass.txt","r")
		pas=f.read()
		pas=pas[0:len(pas)-1]
		try:
			mydb = sql.connect(host="localhost", user="root",passwd=pas,database="Instafood",charset="utf8")
		except:
			mydb = sql.connect(host="localhost", user="root",passwd=pas,charset="utf8")

		mycursor = mydb.cursor()

		# Setting Name and Logo for TK window

		li = ["Chinese", "South Indian", "North Indian", "Desserts", "Beverages", "Fast Food", "Snacks", "Italian"]

		# Radiobuttons for Cusine
		cus = StringVar()
		cus.set("Chinese")
		cusine = ""
		CusineButton = Radiobutton(frame2, text="")

		top = ""

		def Submit():
			global cusine
			global frame3
			cusine = cus.get()
			frame2.destroy()
			frame3 = Frame(win3, width=20, height=430, bg="pale green")
			frame3.place(x=0, y=100)

			SelectRest()

		def SetButtons():
			global submitButton
			global exitButton
			a = 40
			count = 1
			for i in range(0, len(li), 1):
				if (i == int(len(li)) / 2):
					a = 170
					count = 1
				li[i] = li[i]
				Radiobutton(frame2, text=li[i], variable=cus, value=li[i], padx=10, bg="pale green").place(x=a, y=(
																													  count) * 50)
				count += 1

			Button(frame2, text="Submit", command=Submit, padx=10, pady=10, bg="orange").place(x=350, y=80)
			Button(frame2, text="Exit", padx=15, pady=10, command=frame2.quit, bg="orange").place(x=350, y=170)

		global cart
		global Restaurant

		restli = ["None"]
		restaurant = ""
		option = StringVar()

		SetButtons()

		def xlsxfile():
			global cart

			workbook = xlsxwriter.Workbook('data.xlsx')
			worksheet = workbook.add_worksheet()
			worksheet.write('A1', 'FoodName') 
			worksheet.write('B1', 'Quantity') 
			worksheet.write('C1', 'Price') 
			li=list(cart.values())

			for i in range(0,len(li)):
				worksheet.write("A"+str(i+2),li[i][0])
				worksheet.write("B"+str(i+2),li[i][2])
				worksheet.write("C"+str(i+2),li[i][1])
			workbook.close()

		def tot():
			global cart
			GrandTotal=0
			for j in cart:
				li = cart[j]
				qty = li[2]
				price = li[1]
				if qty != 0:
					GrandTotal += qty * price
			return GrandTotal	

		def mklist(n):
			for i in range(1,n+1):
				yield[]    
		
		

		def Bill():
			r=Tk()
			r.withdraw()
			messagebox.showinfo("Success","The Bill has been Generated. Please check the folder. Thank You and Order Again :)")
			r.destroy()

			xlsxfile()

			global maxrow
			global maxcolumn
			global lists

			pdfmetrics.registerFont(TTFont('VeraIt','VeraIt.ttf'))
			wb=openpyxl.load_workbook("data.xlsx") 
			sheet=wb.active
			maxrow=sheet.max_row
			maxcolumn=sheet.max_column
			c = canvas.Canvas('Bill.pdf', pagesize=A4)
			width, height = A4
			maxrow=sheet.max_row
			maxcolumn=sheet.max_column
			elements = []
			lists=list(mklist(maxrow))
			total=tot()

			im=Image.open("instafood.png")
			c.drawInlineImage(im,455,600)
		  
		  
			for i in range(1,maxrow+1):
				for j in range(1,maxcolumn+1):
					cellobj=sheet.cell(row=i,column=j)
					if j<=3:
						lists[i-1].append(str(cellobj.value))
					else:
						pass
			t=Table(lists)
			t.setStyle(TableStyle([
							 ('GRID',(0,0),(-1,-1),0.5,colors.grey),
							 ('BOX',(0,0),(-1,-1),2,colors.black),
							 ]))

			t.wrapOn(c, width, height)
			t.drawOn(c, 180,450)
			c.setLineWidth(2.5)
			c.line(15,15,575,15)
			c.line(575,15,575,815)
			c.line(575,815,15,815)
			c.line(15,815,15,15)
			c.drawString(225,430,"*Item Total:"+str(total))
			c.drawString(225,405,"*GST:"+str((8*total)/100))
			c.drawString(225,380,"*Delivery Charges:"+str(30))
			c.drawString(225,355,"*GrandTotal:"+str(((8*total)/100)+total+30))
			c.drawString(100,650,"Thanks for choosing Intsafood:),Bon Appetit!!!! ") 
			c.drawString(100,638,"Here are your order details:")  
			styles = getSampleStyleSheet()
			c.setFont("Courier",25)
			c.drawString(100,750,"INVOICE")  
			t3 = "Disclaimer:This is an acknowledgement of Delivery of Order and not an actual invoice.Details mentioned above including the menu prices and taxes(as applicable) are provided by the restaurant to Instafood.Resposiblity of charging taxes lies with the Restaurant and Instafood disclaims any liablity that may arise in this respect."
			p3 = Paragraph(t3, style=styles["Normal"])
			p3.wrapOn(c,500,200)
			p3.drawOn(c, 50, 50)
			c.setFont("Courier",8)
			c.drawString(50,35,"*All prices are of the value Indian Ruppees(INR)")
			
			c.save()


		def SelectRest():
			global restli
			global cusine
			global veg
			global option

			veg = IntVar()
			option = StringVar()

			def SubmitRest():
				global currentpos
				global restaurant
				global frame4
				currentpos = 0
				restaurant = option.get()
				SelectFood()

			mycursor.execute("SELECT DISTINCT(Restaurant) from food where CusineOfFood=(%s);", (cusine,))
			record = mycursor.fetchall()

			restli = ["Any"]

			for i in record:
				restli.append(i[0])
			option.set("Any")

			Label(frame3, text="\t\t\t", bg="pale green").grid(row=0, column=0)
			Label(frame3, text="\t\t\t", bg="pale green").grid(row=0, column=1)
			Label(frame3, text="\t\t", bg="pale green").grid(row=0, column=2)

			Label(frame3, text="Restaurant: ", bg="pale green").grid(row=1, column=0, padx=5, pady=5, sticky=E)

			op = OptionMenu(frame3, option, *restli)
			op.grid(row=1, column=1, padx=5, pady=5, sticky=W)
			op.config(bg="pale green")

			cb = Checkbutton(frame3, text="Veg Only", variable=veg, bg="pale green")
			cb.grid(row=3, column=0, padx=7, pady=7, sticky=E)

			Button(frame3, text="Submit", command=SubmitRest, bg="orange").grid(row=4, column=0, columnspan=5, padx=7,
																				pady=7, sticky=W + E)
			Button(frame3, text="Back", command=lambda: back(frame2), bg="orange").grid(row=5, column=0, columnspan=5,
																						padx=7, pady=7, sticky=W + E)

		def SelectFood():
			global cusine
			global restaurant
			global currentpos
			global start
			global cart
			global veg
			global frame4
			global start

			currentpos = 0
			frame3.destroy()
			frame4 = Frame(win3, width=20, height=400, bg="pale green")
			frame4.place(x=0, y=100)

			if (restaurant == "Any" and veg.get() == 1):
				mycursor.execute("SELECT FoodName,Price,Restaurant from food where CusineOfFood=(%s) and Veg=1",
								 (cusine,))

			elif (restaurant == "Any" and veg.get() == 0):
				mycursor.execute("SELECT FoodName,Price,Restaurant from food where CusineOfFood=(%s)", (cusine,))
			elif (veg.get() == 1):
				mycursor.execute(
					"SELECT FoodName,Price,Restaurant from food where CusineOfFood=(%s) and Restaurant=(%s) and Veg=1",
					(cusine, restaurant))
			else:
				mycursor.execute(
					"SELECT FoodName,Price,Restaurant from food where CusineOfFood=(%s) and Restaurant=(%s) and Veg=1",
					(cusine, restaurant))

			record = mycursor.fetchall()

			qty = IntVar()
			qty.set(0)

			def change(n):
				global start
				start = start + n
				frame4.destroy()
				SelectFood()

			def Buttondisplay(num):
				global cusine

				if (num - start == 0):
					Button(frame4, text="Add To Cart", bg="orange",
						   command=lambda: Addtocart(e1.get(), record[num][0], record[num][1], record[num][2], cusine)).grid(
						row=0,
						column=4,
						padx=5,
						pady=5)
				elif (num - start == 1):
					Button(frame4, text="Add To Cart", bg="orange",
						   command=lambda: Addtocart(e2.get(), record[num][0], record[num][1], record[num][2], cusine)).grid(
						row=1,
						column=4,
						padx=5,
						pady=5)
				elif (num - start == 2):
					Button(frame4, text="Add To Cart", bg="orange",
						   command=lambda: Addtocart(e3.get(), record[num][0], record[num][1], record[num][2], cusine)).grid(
						row=2,
						column=4,
						padx=5,
						pady=5)
				elif (num - start == 3):
					Button(frame4, text="Add To Cart", bg="orange",
						   command=lambda: Addtocart(e4.get(), record[num][0], record[num][1], record[num][2], cusine)).grid(
						row=3,
						column=4,
						padx=5,
						pady=5)
				else:
					Button(frame4, text="Add To Cart", bg="orange",
						   command=lambda: Addtocart(e5.get(), record[num][0], record[num][1], record[num][2], cusine)).grid(
						row=4,
						column=4,
						padx=5,
						pady=5)

			def Addtocart(qty, FoodName, price, rest, cus):
				global currentpos
				global cart

				self.addorder([FoodName, rest,price,qty,cus])
				try:
					l.destroy()
				except:
					pass

				try:
					l = Label(frame4, text="\t\t", bg="pale green")
					qty = int(qty)
					r = len(rest)
					f = len(FoodName)

					FoodId = cusine[0:3] + FoodName[0:2] + FoodName[f - 2:f] + rest[0:2] + rest[r - 2:r]
					if (qty <= 7):
						cart[FoodId] = (FoodName, price, qty)
						
					else:
						l = Label(frame4, text="MAX 7 items.", bg="pale green")
						l.grid(row=8, column=0)

				except:
					l = Label(frame4, text="Enter an integer value please.", bg="pale green")
					l.grid(row=8, column=0)

			if (len(record) >= start + 5):
				end = start + 5
				nex = Button(frame4, text=">>Next", command=lambda: change(+5), bg="orange")
			else:
				end = len(record)
				nex = Button(frame4, text=">>Next", command=lambda: change(+5), state=DISABLED, bg="orange")
			if (start == 0):
				back = Button(frame4, text="<<Back", command=lambda: change(-5), state=DISABLED, bg="orange")
			else:
				back = Button(frame4, text="<<Back", command=lambda: change(-5), bg="orange")

			e1 = Entry(frame4, width=5, borderwidth=3)
			e2 = Entry(frame4, width=5, borderwidth=3)
			e3 = Entry(frame4, width=5, borderwidth=3)
			e4 = Entry(frame4, width=5, borderwidth=3)
			e5 = Entry(frame4, width=5, borderwidth=3)

			EntryList = [e1, e2, e3, e4, e5]

			for j in range(start, end):
				Label(frame4, text=record[j][0], bg="pale green").grid(row=currentpos, column=0, padx=5, pady=3,
																	   sticky=W)
				Label(frame4, text=record[j][1], bg="pale green").grid(row=currentpos, column=1, padx=10, pady=3)
				Label(frame4, text="QTY:", bg="pale green").grid(row=currentpos, column=2, pady=3)
				EntryList[j - start].grid(row=currentpos, column=3, pady=3)
				EntryList[j - start].insert(0, "0")
				Buttondisplay(j)
				currentpos += 1

			def go():
				frame4.destroy()
				self.order_page()

			nex.grid(row=6, column=4, columnspan=2, padx=10, pady=10)
			back.grid(row=6, column=0, columnspan=2, padx=10, pady=10)
			Button(frame4, text="Cusines", command=go, bg="orange").grid(row=6, column=2, columnspan=2, padx=10,
																		 pady=10)
			Button(frame4, text="Generate Bill", command=Bill, bg="orange").grid(row=7, column=0, columnspan=5,
																					 padx=7, pady=3, sticky=W + E)

		def back(dest):
			if (dest == frame2):
				frame3.destroy()
				self.order_page()

		frame2.mainloop()


class Transaction():
	def tracker(self):

		screen = turtle.Screen()
		screen.setup(1366, 768)
		screen.screensize(400, 400)
		screen.bgpic('map.png')  
		k = 3
		for i in range(0, k):
			a = turtle.Turtle()
			a.hideturtle()
			a.penup()
			a.shapesize(3, 3, 3)
			a.color("black", "red")
			a.speed(i + 1)
			a.pensize(8)
			a.goto(-40, 350)
			a.write("Restaurant", font=("Comic Sans MS", 16, 'normal', 'bold', 'italic', 'underline'))
			a.showturtle()
			a.pendown()
			a.fd(283)
			a.goto(243, 70)
			a.rt(180)
			a.fd(665)
			a.lt(90)
			a.fd(143)
			a.lt(90)
			a.fd(130)
			a.rt(90)
			a.fd(100)
			a.lt(90)
			a.write("Your Location", font=("Comic Sans MS", 16, 'normal', 'bold', 'italic', 'underline'))
		a.clearscreen()
		a.bye()

	def time_calc(self):
		root = tk.Tk()
		root.withdraw()
		root.iconbitmap('Logo.ico')

		dis = askstring("Time Estimator", "Enter the approximate distance from the Restaurant to your Location(in KM):")
		
		#Since we are not able to simulate life time situation we randomly choose the traffic
		n=random.randint(0,2)
		traff=["Heavy","Average","Light"]

		r=Tk()
		r.withdraw()	
		s="Looks like your area has "+traff[n]+" traffic."
		messagebox.showinfo("Delivery",s)
		r.destroy()

		#tra = askstring("Time Estimator", "traffic in your location=H-heavy/M-medium/L-little:")
		d = int(dis)
		if n == 0:
			t = int((d / 20) * 60)
		elif n == 1:
			t = int((d / 35) * 60)
		elif n == 2:
			t = int((d / 50) * 60)

		r=Tk()
		r.withdraw()	
		s="Approximate time of Arrival ", t, "minutes"
		messagebox.showinfo("Delivery",s)
		r.destroy()

#       screen.update()
#        time.sleep(2)
#       screen.bgpic('C:\\Users\\harsh\\OneDrive\\Documents\\ACHINTYA!\\COMPUTERS\\Class 12\\order again.png')  

count = 1
backcount = 0
accounts = []

client = MongoClient("mongodb+srv://user1:alpine2003@cluster0-qazji.gcp.mongodb.net/<dbname>?retryWrites=true&w=majority")
db = client.get_database('accounts_db')
records = db.account_records

log = Login()
home = Homepage()
trans = Transaction()

def main():
	log.sign_in_check()
	global accounts
	accounts = list(map(lambda item: list(item.values()), records.find()))
	r=Tk()
	r.withdraw()	
	messagebox.showinfo("Pop-Up","You have sucessfully logged in.")
	r.destroy()
	home.order_page()
	home.FoodData()

main()
